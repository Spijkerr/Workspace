# -*- coding: utf-8 -*-
"""
Created on Mon Jan 18 14:39:05 2016

@author: OffermanTW1
"""




import pandas as pd
import openpyxl as xl 
from tkinter import filedialog
import tkinter as tk
import re
import random
from datetime import datetime

class DataCutter():
    def __init__(self, protect_cols = None, protect_rows = None, col_limit = 0.05, row_limit = 0.50):
        self.protected_rows = protect_rows
        self.protected_cols = protect_cols
        self.col_limit = col_limit
        self.row_limit = row_limit 
		
    def cut_redundant_data(self, data):
        #First delete all irrelevant columns

        col_count = {}
        for row in range(len(data)):
            for col in range(len(data[row])):
                if data[row][col] != None and data[row][col] != "":
                    col_count[col] = col_count[col] + 1 if col in col_count else 1
                else:
                    col_count[col] = col_count[col] if col in col_count else 0 

            
        total_vals = [col_count[col] for col in col_count]
        total_vals = sum(total_vals)
        average_col = total_vals/len(col_count)
        cols_to_delete = [col for col in col_count if col_count[col] < average_col * self.col_limit]

        if self.protected_cols:
            cols_to_delete = [col for col in cols_to_delete if col not in cls.protected_cols]
        
        row_count = {}
        for row in range(len(data)):
            new_row = [data[row][col] for col in range(len(data[row])) if col not in cols_to_delete]
            data[row] = new_row
            
            if len(new_row) == 0:
                row_count[row] = 0
            for col in range(len(data[row])):
                if data[row][col]:
                    row_count[row] = row_count[row] + 1 if row in row_count else 1
                else: 
                    row_count[row] = row_count[row] if row in row_count else 0
                    
        
        total_vals = [row_count[row] for row in row_count]
        total_vals = sum(total_vals)
        """
        Wouldn't this work as total_vals = sum([row_count[row] for row in row_count])?
        """
        
        average_row = total_vals/len(row_count)

        rows_to_delete = [row for row in row_count if row_count[row] < (average_row * self.row_limit)]
        
        rows_to_delete = sorted(rows_to_delete, reverse = True)
        

        if self.protected_rows: 
            rows_to_delete = [row for row in rows_to_delete if row not in cls.protected_rows]
        for row in rows_to_delete:
            data.pop(row)
            
        return data


class DtypeDetector():
    def __init__(self, assign_dtype = None):
        if assign_dtype:
            assert type(assign_dtype) == dict, "Assign_dtype needs to be dict"
            
        self.assign_dtype = assign_dtype if assign_dtype else {}

            
    def transform_dtypes_in_list(self, data):
        total_header = [header.strip() for header in data[0]]
        dtype_dict = self.detect_dtypes(data)
        
        for key in self.assign_dtype:
            dtype_dict[key] = self.assign_dtype[key]
        
        for row in range(len(data)):
            for col in range(len(data[0])):
                
                header = total_header[col]
                
                dtype = dtype_dict.get(header, None)
                data[row][col] = self.convert_to_dtype(data[row][col], dtype)

        return data 

    def detect_dtypes(self, data):
        header = data[0]
        
        dtype_to_regex = {'ddmmyyyy_slash': r"([0-2]?[0-9]|3[0-1])//(0[0-9]|1[0-2])//[1-2][0-9][0-9][0-9]",
        'ddmmyyyy_colon': r"([0-2]?[0-9]|3[0-1]):(0[0-9]|1[0-2]):[1-2][0-9][0-9][0-9]",
        'ddmmyyyy_dot': r"([0-2]?[0-9]|3[0-1])\.(0[0-9]|1[0-2])\.[1-2][0-9][0-9][0-9]",
        'ddmmyyyy_dash': r"([0-2]?[0-9]|3[0-1])-(0[0-9]|1[0-2])-[1-2][0-9][0-9][0-9]",
        'float_dotdecimal': r"^-?(([0-9]?[0-9]?(,?[0-9][0-9][0-9])*\.[0-9]?[0-9]?)|([0-9]?[0-9]?))$",
        'float_commadecimal': r"^-?(([0-9]?[0-9]?(.?[0-9][0-9][0-9])*,[0-9]?[0-9]?)|([0-9]?[0-9]?))$"}

        """
        Would the regex r"[0-3][0-9]/[0-1][0-9]/[1-2][0-9][0-9][0-9]" 
        not produce the same results when it comes to ddmmyyyy_backslash?
		No, the ? after [0-2] is to make it optional if the zero drops. 
        """
        
        #Take a sample of integers. These are the indexes we will use for validating the dtype
        #We will use n integers, with a maximum of 100.


        sample = [random.randint(1,len(data)-1) for n in range(0, min(len(data),100))]
        
        header_to_dtype_map = {}
        for col in range(len(header)):
            for regex in dtype_to_regex:

                for element in sample:
                    x = data[element][col]
                    x = x.strip() if type(x) == str else x 
                    match = None 
                    
                    if type(x) == str: 
                        match = re.match(dtype_to_regex[regex], x)

                    if x == None:
                        continue
                    
                    elif match == None:
                        break
                else:
                    header_to_dtype_map[header[col].strip()] = regex
                    break
                
        return header_to_dtype_map
                
    def convert_to_dtype(self, data, dtype = None):
        if dtype == None:            
            return str(data)
            
        if dtype == 'float_dotdecimal':
            return self.float_dotdecimal(data)

        elif dtype == 'float_commadecimal':
            return self.float_commadecimal(data)
            
        elif dtype == 'ddmmyyyy_dot':
            return datetime.strptime(data, '%d.%m.%Y')
        
        elif dtype == 'ddmmyyyy_slash':
            return datetime.strptime(data, '%d/%m/%Y')
        
        elif dtype == 'ddmmyyyy_dash':
            return datetime.strptime(data, '%d-%m-%Y')
        
        elif dtype == 'ddmmyyyy_colon':
            return datetime.strptime(data, '%d:%m:%Y')  
        else:
            raise ValueError('Unknown dtype: ' + str(dtype))
        
    def float_commadecimal(self, data):
        #Return a flaot or None if a None-value is passed
        if data:
            data = data.replace('.', '')
            data = data.replace(',', '.')
            return float(data.strip())
        else:
            return data
    """
    Should these not be the other way around? ',' to '.' first, then '.' to ''. 
    """
            
    def float_dotdecimal(self, data):
        #Return a flaot or None if a None-value is passed
        if data:
            data = data.replace(',', '')
            return float(data.strip())
        else:
            return data
        
    
class DataListLoader():
    @classmethod
    def find_extension_and_load(self, data_location):
        #select the extension by selecting the last element after splitting
        lower_extension, old_extension = self.find_extension(data_location)
        data_location = data_location.replace(old_extension, lower_extension)
        
        if lower_extension in ['xlsx', 'xls']:
            data = self.load_excel_to_lists(data_location)
        elif lower_extension in ['txt']:
            data = self.load_txt_to_lists(data_location)
        elif lower_extension in ['csv']:
            raise NameError('CSV not yet supported')
        else:
            raise NameError('Extension of file is not recognized')
            
        return data
        
    @classmethod
    def find_extension(self, data_location):
        original_extension = data_location.split('.')[-1]
        lowercase_extension = original_extension.lower()
        return lowercase_extension, original_extension
          
    @classmethod          
    def load_txt_to_lists(self, filepath):

        
        loaded_list = open(filepath) #open the file 
        loaded_list = loaded_list.read() #read the file
        loaded_list = loaded_list.split('\n') #split for line items
        loaded_list = [line.split('\t') for line in loaded_list] #split on tabs for tab delimited.
        
        return loaded_list
    
    @classmethod
    def load_excel_to_lists(self, filepath):
        workbook = xl.load_workbook(filepath)
        loaded_list = []
        sheet = workbook.active
        for row in sheet.rows:
            row_list = []
            for col in row:
                row_list.append(col.value)
            
            loaded_list.append(row_list)
            
        return loaded_list
		

class DataFrameLoader():
    def load_data(self, data_location = None,
                  cut_redundant_data = True,
                  dtype_detection = True,
                  protect_rows = None, 
                  protect_cols = None,
                  col_is_index = None,
                  assign_dtype = None):
        """
        This function consists of the following steps:
        1. Loading the data, from various possible sources
        2. Filtering the redundant data
        3. Turning the data into a Pandas DataFrame
        """                     
                     
 
        """1. Loading the data"""

        
        #If data_location is already a DataFrame, return it. 
        if isinstance(data_location, pd.DataFrame):
            return data_location
            
        #If there is no data location, open a dialog box to select a file
        elif data_location == None:
            root = tk.Tk()
            root.withdraw()
            data_location = filedialog.askopenfilename()
            
        #If data_location is a list, that list is the data.
        if isinstance(data_location, list):
            data = data_location
            
        #If data_location is a string, laod it with the appropriate extension
        elif isinstance(data_location, str):
            data = DataListLoader.find_extension_and_load(data_location)


 
        """2. Filtering the redundant data and detecting dtypes"""
        if cut_redundant_data:
            cutter = DataCutter(protect_cols, protect_rows)
            data = cutter.cut_redundant_data(data)

        if dtype_detection:
            detector = DtypeDetector(assign_dtype)
            data = detector.transform_dtypes_in_list(data)
            
            
        """3. Turning the data in a DataFrame"""

        total_header = [header.strip() for header in data[0]] 
        data = pd.DataFrame(data[1:], columns = total_header) 

        if col_is_index:
            data = data.set_index(col_is_index)
            total_header.remove(col_is_index)
        else:
            data.index.names = ['ID']
        
        total_header = [header for header in total_header if header != '']   
        data = data[total_header] #re-order the dataframe to resemble the list-order
            
        return data


#Write functions 	
	
class DataWriter():
    @classmethod
    def find_extension_and_open(self, data_location):
        extension = self.find_extension(data_location)
        
        if extension in ['xlsx', 'xls']:
            wb = xl.Workbook()
            wb.save(data_location)
            
        elif extension in ['txt']:
             open(data_location, 'w')
        elif extension in ['csv']:
            raise NameError('CSV not yet supported')
        else:
            raise NameError('Extension of file is not recognized')	
            
    @classmethod
    def find_extension_and_append(self, location, data_list):
        extension = self.find_extension(location)
        
        if extension in ['xlsx', 'xls']:
            self.append_list_to_excelfile(location, data_list)
        elif extension in ['txt']:
            self.append_list_to_txtfile(location, data_list)
        else:
            raise NameError('Extension of file is not yet recognized')      
     
    @classmethod 
    def append_list_to_excelfile(self, location, datalist):
        wb = xl.load_workbook(location)
        ws = wb.active
        for row in datalist:
            new_row = ws.max_row
            for col in range(len(row)):
                ws.cell(column = col+1, row = new_row + 1, value=row[col])            
        wb.save(location)

    def append_list_to_txtfile(self, location, datalist):
        file = open(location, 'a')
        
        for row in datalist:
            string = ""
            for col in row:
                string += str(col) + "\t"
            string += "\n"
            file.write(string)
            
        file.close()

#print (DataFrameLoader.load_data(data_location = None))