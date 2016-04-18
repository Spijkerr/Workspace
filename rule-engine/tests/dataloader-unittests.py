# -*- coding: utf-8 -*-
"""
Created on Mon Apr  4 10:25:52 2016

@author: OffermanTW1
"""

#Data Loader test cases 

import DataLoader3 as DL
import unittest
from pandas import DataFrame
from pandas.util.testing import assert_frame_equal
import openpyxl as xl
#
#if sys.executable != 'J:\Applications\BB-Python\Python\python.exe':
#    raise Exception ('Python is run from the wrong location')

def locator(filename):
    abs_file_path = "J:/Applications/BB-Python/Python/Testcases/" + filename
    return abs_file_path


class DataListLoaderTestCases(unittest.TestCase):
    
    """
    Create a test for empty headers. Thiw will, however, require the overhaul of quite a few tests.
    """
    dataloader = DL.DataListLoader()
    def test_datalistloader1(self):
        #Test 1: Check whether data is loaded properly
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[0:2], [['Header1', 'Header2', 'Header3'],
                                  ['Regel1', 'Regel2', 'Regel3']])
                                       
    def test_datalistloader2(self):
        #Test 2: Check whether capitalized extension is still recognized. 
        result = self.dataloader.find_extension_and_load(locator('datalistloader2.TXT'))
        
        self.assertEqual(result, [['Header1', 'Header2', 'Header3'],
                                  ['Regel1', 'Regel2', 'Regel3']])        
                                  
    def test_datalistloader3(self):
        #Test 3: Empty dataset 
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[3:8], [['','','',''],['','','',''],
                                  ['','','',''],['','','',''],['','','','']])   
                                 
    def test_datalistloader4(self):
        #Test 4: Check whether special characters are properly handled
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[8], ['\'"%Regel7', "'Regel8#'", '&"Regel9"\''])   
                                   
    def test_datalistloader5(self):
        #Test 5: check whether empty rows and columns influence the data in between
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[7:10], [['', '', '', ''], ['\'"%Regel7', "'Regel8#'", '&"Regel9"\''], ['', '', '', '']])           
                                                          
                                  
    def test_datalistloader6(self):
        #Test 6: Check whether data is loaded properly
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
        
        self.assertEqual(result[0:3], [['Header1', 'Header2', 'Header3'], ['Regel1', 'Regel2', 'Regel3'], 
                                      ['"3454235', '&85674#', "'sadaf'"]])
                                       
    def test_datalistloader7(self):
        #Test 7: Check whether capitalized extension is still recognized. 
        result = self.dataloader.find_extension_and_load(locator('datalistloader2.XLSX'))
        
        self.assertEqual(result, [['Header1', 'Header2', 'Header3'],
                                  ['Regel1', 'Regel2', 'Regel3']])        
                                                                                                           
    def test_datalistloader8(self):
        #Test 8: Check how empty cells are handled 
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
        
        self.assertEqual(result[3:8], [[None, None, None],[None, None, None],[None, None, None],
                                       [None, None, None],[None, None, None]])                                          
                                                                            
    def test_datalistloader9(self):
        #Test 9: Check empty file
        result = self.dataloader.find_extension_and_load(locator('datalistloader3.xlsx'))
        
        self.assertEqual(result, [[]])                              
                                  
#    def test_datalistloader10(self):
#        """
#        Seems to be an issue with my python formula, not the script
#        """
#        #Test 10: Check whether csv is not supported
#        self.assertRaises(NameError('CSV not yet supported'), self.dataloader.find_extension_and_load(locator('datalistloader1.csv')))       

#    def test_datalistloader11(self):
#        """
#        Seems to be an issue with my python formula, not the script
#        """
#        #Test 11: Check whether cvs is not supported
#        self.assertRaises(NameError('Extension of file is not recognized'), self.dataloader.find_extension_and_load(locator('datalistloader1.cvs')))    
#        
    def test_datalistloader12(self):
        #Test 12: Check loading with an empty first cell
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[10], ['','Regel10','', ''])        
                      
    def test_datalistloader13(self):
        #Test 13: Check loading with only a filled first cell
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
        
        self.assertEqual(result[11], ['Regel11', '', '' , ''])                              
        
#    def test_datalistloader14(self):
#        """
#        Required? It would only matter when manually creating columns in a .txt. Still, it might be used at some point. 
#        Currently no support for changing column width as 1 invalid column screws up the entire datalist
#        """
#        #Test 14: Check how changing column widths are handled
#        result = self.dataloader.find_extension_and_load(locator('datalistloader1.txt'))
#        
#        self.assertEqual(result[13:15], [])       

    def test_datalistloader15(self):
        #Test 15: Checks special chartacters splt by an empty cell
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
        
        self.assertEqual(result[12], ['""Regel5', None, '&"Regel6\''])  
                                       
#    def test_datalistloader16(self):
#        """
#        Right now returns the formula. Look for a way to return only the value
#        """
#        #Test 16: Check whether formulas return values rather than formulas
#        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
#        
#        self.assertEqual(result[8], [18, None, None])       
                                       
    def test_datalistloader17(self):
        #Test 16: Check whether quoted formulas retain their value as such
        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
        
        self.assertEqual(result[9], [None, '=17^6', "Regel4#'"])                                            
                                       
                                       
#    def test_datalistloader18(self):
#        """
#        Currency from Dutch to National usage is working as intended. However when importing a number using national notation
#        the number is handled as a string rather than an integer. input check when a comma is found to see whether the length is
#        3 or more. If so, remove comma and pass on the edited number. In case this module might also be used for analysis on 
#        more than economic data, the option whether to use this function should have a toggle function.
#        """
#        #Test 18: Check how formulas are handled
#        result = self.dataloader.find_extension_and_load(locator('datalistloader1.xlsx'))
#        
#        self.assertEqual(result[9], [147512.25, 147512.25, 147512.25])           
                
    
class DataCutterTestCases(unittest.TestCase):
    datacutter = DL.DataCutter(protect_cols = None, protect_rows = None, col_limit = 0.05, row_limit = 0.50)
    """
    Input check to see whether kwargs datacutter = kwargs DL.Datacutter
    """
    
    def test_datacutter1(self):
        #Test 1: Check whether the datacutter cuts rows properly
        dataset = [['Header1', 'Header2'], 
                   ['',''],
                   ['Data1', 'Data2']]
        
        cut_data = self.datacutter.cut_redundant_data(dataset)
        self.assertEqual(cut_data,[['Header1', 'Header2'], ['Data1', 'Data2']] )
        
    def test_datacutter2(self): 
        #Test 2: Check whether the datacutter cuts columns properly
        dataset = [['Header1', '', 'Header2'], 
                   ['', '',''],
                   ['Regel1', '',  'Regel2'], 
                   ['', '','']]
        
        cut_data = self.datacutter.cut_redundant_data(dataset)
        
        self.assertEqual(cut_data,[['Header1', 'Header2'], ['Regel1', 'Regel2']] )
        
    def test_datacutter3(self): 
        #Test 3: Check whether the protected rows are not cut
        datacutter = DL.DataCutter(protect_cols = 'test' , protect_rows = None)
        dataset = [['Header1', 'Header2', 'Header3'], 
                   ['', '',''],
                   ['Regel1', '', 'Regel2'], 
                   ['', '','']]
        
        cut_data = datacutter.cut_redundant_data(dataset)
        
        self.assertEqual(cut_data,[['Header1', 'Header2', 'Header3'], ['Regel1', '', 'Regel2']])      
        
        
    """
    function incorrect. How to call the proper row to be protected?
    """   
#    def test_datacutter4(self): 
#        #Test 4: Check whether the protected columns not cut
#        datacutter = DL.DataCutter(protect_cols = None , protect_rows = 'test', col_limit = 0.05, row_limit = 0.50)
#        dataset = [['Header1', '', 'Header2'], 
#                   ['', '',''],
#                   ['Regel1', '', 'Regel2'], 
#                   ['test', '','']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data,[['Header1', 'Header2'], ['Regel1', 'Regel2'], ['test', '', '']] )    
     
     
    """
    For the following tests to work, the row/column filtering needs to be re-evaluated. Empty rows and cells should be removed
    first before calculating the averages. 
    """
    """
    Include a check so that when col_count[col] = 1, deletion follows, as that should be only the
    header, and another removal of empty rows after the column deletion.
    """
    
    def test_datacutter5(self): 
        #Test 5: Check whether no restrions on rows or colums returns the whole dataset
        datacutter = DL.DataCutter(col_limit = 0.33, row_limit = 0)
        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
                   ['Regel3', 'Regel7', '', '', '', ''],
                   ['Regel4', '', '', '', '', ''],
                   ['', '', '' ,'' ,'' ,'']]
        
        cut_data = datacutter.cut_redundant_data(dataset)
        
        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5'], 
                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10', ''],
                                    ['Regel2', 'Regel6', 'Regel9', '', ''], 
                                    ['Regel3', 'Regel7', '', '', ''],
                                    ['Regel4', '', '', '', ''],
                                    ['', '', '', '', '']])
                                    
#    def test_datacutter6(self): 
#        #Test 6: Check whether column limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0.34, row_limit = 0)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4'], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10'],
#                                    ['Regel2', 'Regel6', 'Regel9', ''], 
#                                    ['Regel3', 'Regel7', '', ''],
#                                    ['Regel4', '', '', ''],
#                                    ['', '', '', '']])                                 
#                                   
#    def test_datacutter7(self): 
#        #Test 7: Check whether column limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0.67, row_limit = 0)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3'], 
#                                    ['Regel1', 'Regel5', 'Regel8'],
#                                    ['Regel2', 'Regel6', 'Regel9'], 
#                                    ['Regel3', 'Regel7', ''],
#                                    ['Regel4', '', ''],
#                                    ['', '', '']])  
#                                   
#    def test_datacutter8(self): 
#        #Test 8 : Check whether column limits work correctly
#        datacutter = DL.DataCutter(col_limit = 1.01, row_limit = 0)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2'], 
#                                    ['Regel1', 'Regel5'],
#                                    ['Regel2', 'Regel6'], 
#                                    ['Regel3', 'Regel7'],
#                                    ['Regel4', ''],
#                                    ['', '']])   
#
#    def test_datacutter9(self): 
#        #Test 9: Check whether column limits work correctly
#        datacutter = DL.DataCutter(col_limit = 1.34, row_limit = 0)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1'], 
#                                    ['Regel1'],
#                                    ['Regel2'], 
#                                    ['Regel3'],
#                                    ['Regel4'],
#                                    ['']])  
#                                    
#    def test_datacutter10(self): 
#        #Test 10: Check whether column limits work correctly
#        datacutter = DL.DataCutter(col_limit = 1.67, row_limit = 0)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [[], [], [], [], [], []])   
#
#    def test_datacutter11(self): 
#        #Test 11: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 0.33)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5', ''], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                                    ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                                    ['Regel3', 'Regel7', '', '', '', ''],
#                                    ['Regel4', '', '', '', '', '']])                                    
#                                   
#    def test_datacutter12(self): 
#        #Test 12: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 0.34)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5', ''], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                                    ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                                    ['Regel3', 'Regel7', '', '', '', '']])                                   
#                                   
#    def test_datacutter13(self): 
#        #Test 13: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 0.68)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5', ''], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                                    ['Regel2', 'Regel6', 'Regel9', '', '', '']])                                      
#                                   
#    def test_datacutter14(self): 
#        #Test 14: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 1.01)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5', ''], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', '']])                                     
#                                   
#    def test_datacutter15(self): 
#        #Test 15: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 1.34)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4', 'Header5', '']])                                      
#                                   
#    def test_datacutter16(self): 
#        #Test 16: Check whether row limits work correctly
#        datacutter = DL.DataCutter(col_limit = 0, row_limit = 1.67)
#        dataset = [['Header1', 'Header2', 'Header3', 'Header4', 'Header5',''], 
#                   ['Regel1', 'Regel5', 'Regel8', 'Regel10', '', ''],
#                   ['Regel2', 'Regel6', 'Regel9', '', '', ''], 
#                   ['Regel3', 'Regel7', '', '', '', ''],
#                   ['Regel4', '', '', '', '', ''],
#                   ['', '', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        
#        self.assertEqual(cut_data, [])                                     
#                                   
#    def test_datacutter17(self): 
#        #Test 17: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 0.33, row_limit = 0.33)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4','Header5'], 
#                                    ['Regel1', 'Regel5', '', 'Regel8', 'Regel9'],
#                                    ['Regel2', 'Regel6', '', '', 'Regel10'], 
#                                    ['Regel3', '', '', '', ''],
#                                    ['Regel4', 'Regel7', '' ,'' ,'']])                                             
#                                   
#    def test_datacutter18(self): 
#        #Test 18: Check whether the combination of row and column limits works correctly
#
#        datacutter = DL.DataCutter(col_limit = 0.33, row_limit = 0.67)
#        dataset = [['Header1', 'Header2','',  'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '',  'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'', '']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header3', 'Header4','Header5'], 
#                                    ['Regel1', 'Regel5', '', 'Regel8', 'Regel9'],
#                                    ['Regel2', 'Regel6', '', '', 'Regel10'], 
#                                    ['Regel4', 'Regel7', '' ,'' ,'']])                
#                                    
#    def test_datacutter19(self): 
#        #Test 19: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 0.67, row_limit = 0.33)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [['Header1', 'Header2','Header5'], 
#                                    ['Regel1', 'Regel5', 'Regel9'],
#                                    ['Regel2', 'Regel6', 'Regel10'], 
#                                    ['Regel3', '', ''],
#                                    ['Regel4', 'Regel7','']])       
#    def test_datacutter20(self): 
#        #Test 20: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 1.01, row_limit = 0.56)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [['Header1', 'Header2'], 
#                                    ['Regel1', 'Regel5'],
#                                    ['Regel2', 'Regel6'],
#                                    ['Regel4', 'Regel7']]) 
#
#    def test_datacutter21(self): 
#        #Test 21: Check whether the combination of row and column limits works correctly. 
#        datacutter = DL.DataCutter(col_limit = 0.67, row_limit = 0.42)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [['Header1', 'Header2','Header5'], 
#                                    ['Regel1', 'Regel5', 'Regel9'],
#                                    ['Regel2', 'Regel6', 'Regel10'],
#                                    ['Regel4', 'Regel7', '']])                                       
#
#    def test_datacutter22(self): 
#        #Test 22: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 0.34, row_limit = 0.72)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [['Header1', 'Header2', 'Header4','Header5'], 
#                                    ['Regel1', 'Regel5', 'Regel8', 'Regel9']])      
#                                    
#    def test_datacutter23(self): 
#        #Test 23: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 1.67, row_limit = 3)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [])   
#                                       
#    def test_datacutter24(self): 
#        #Test 24: Check whether the combination of row and column limits works correctly
#        datacutter = DL.DataCutter(col_limit = 0.34, row_limit = 1.43)
#        dataset = [['Header1', 'Header2', '', 'Header3', 'Header4','Header5'], 
#                   ['Regel1', 'Regel5', '', '', 'Regel8', 'Regel9'],
#                   ['Regel2', 'Regel6', '', '', '', 'Regel10'], 
#                   ['', '', '', '', '', ''],
#                   ['Regel3', '', '', '', '', ''],
#                   ['Regel4', 'Regel7', '' ,'' ,'' ,'']]
#        
#        cut_data = datacutter.cut_redundant_data(dataset)
#        self.maxDiff = None
#        self.assertEqual(cut_data, [[], [], [], []])      

class DtypeDetectorTestCase(unittest.TestCase):
    dtypedetector = DL.DtypeDetector()
    
#    def test_dtypedetector1(self):
#        #Test 1: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['Header1', 'Header2', 'Header3'], 
#                ['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        datalist = self.dtypedetector.transform_dtypes_in_list(data) 
#        df = DataFrame(datalist, columns = ['Header1', 'Header2', 'Header3'])
#        print (df)
#        self.assertEqual(df['Header1'][1], '01011990')
#        
#    def test_dtypedetector2(self):
#        #Test 2: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector3(self):
#        #Test 3: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01/31/1990')
#        
#    def test_dtypedetector4(self):
#        #Test 4: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01011990')
#        
#    def test_dtypedetector5(self):
#        #Test 5: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector6(self):
#        #Test 6: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01:31:1990')
#        
#
#    def test_dtypedetector7(self):
#        #Test 7: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01011990')
#        
#    def test_dtypedetector8(self):
#        #Test 8: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector9(self):
#        #Test 9: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01.31.1990')
#        
#
#    def test_dtypedetector10(self):
#        #Test 10: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01011990')
#        
#    def test_dtypedetector11(self):
#        #Test 11: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector12(self):
#        #Test 12: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01-31-1990')
#        
#
#    def test_dtypedetector13(self):
#        #Test 13: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01011990')
#        
#    def test_dtypedetector14(self):
#        #Test 14: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector15(self):
#        #Test 15: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01,31.1990')
#        
#
#    def test_dtypedetector16(self):
#        #Test 16: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01011990')
#        
#    def test_dtypedetector17(self):
#        #Test 17: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '32:01:1990'],
#                ['01.01.1234', '31.12.2999', '32.01.1990'], ['01-01-1234', '31-12-2999', '32-01-1990'],
#                ['01,01.1234', '31,12.2999', '32,01.1990'], ['01.01,1234', '31.12,2999', '32.01,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '31122999')
#        
#    def test_dtypedetector18(self):
#        #Test 18: Check whether the correct dtype is assigned when using slashes as divider
#        data = [['01/01/1000', '31/12/2999', '01/31/1990'], ['01:01:1234', '31:12:2999', '01:31:1990'],
#                ['01.01.1234', '31.12.2999', '01.31.1990'], ['01-01-1234', '31-12-2999', '01-31-1990'],
#                ['01,01.1234', '31,12.2999', '01,31.1990'], ['01.01,1234', '31.12,2999', '01.31,1990']]
#        df = DataFrame(data, columns = ['Header1', 'Header2', 'Header3'])
#        df = self.dtypedetector.transform_dtypes_in_list(data) 
#        df[0]            
#        
#        self.assertEqual(df, '01.31,1990')
#        

                                    
class DataFrameTestCase(unittest.TestCase):
    """
    Expected defaults:
    data_location = None, cut_redundant_data = True, dtype_detection = True, protect_rows = None, 
    protect_cols = None, col_is_index = None, assign_dtype = None
    """

    """
    Device a test to check whether a loaded string gets referencded properly. see Dataloader3.py line 266
    """    
    
    dataframeloader = DL.DataFrameLoader()
    
    
    def test_dataframeloader1(self):
        #Test 1: Checks whether pandaframes handles basic data import correctly
        result = self.dataframeloader.load_data(data_location = locator('dataframeloader1.xlsx'))
      
        t = [['Regel1', 'Regel2', 'Regel3'], ['"3454235', '&85674#', "'sadaf'"]]
        df = DataFrame(t, columns = ['Header1', 'Header2', 'Header3'])    
        df.index.rename('ID', inplace = True)
        assert_frame_equal(result, df)
        
#    def test_dataframeloader2(self):
#        """
#        works as of 05-04, check regularely to ensure proper it keeps working
#        """
#        #Test 2: Checks whether file selection imports data properly
#        result = self.dataframeloader.load_data(data_location = None) #Select dataframeloader1.xlsx
#        
#        t = [['Regel1', 'Regel2', 'Regel3'], ['"3454235', '&85674#', "'sadaf'"]]
#        df = DataFrame(t, columns = ['Header1', 'Header2', 'Header3'])    
#        df.index.rename('ID', inplace = True)
#        assert_frame_equal(result, df)

    def test_dataframeloader3(self):
        #Test 3: Checks whether slicing works correctly when cutting redundant data has been set to false
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'), cut_redundant_data = False)
        
        t = '""Regel5'
        result = result['Header1'].loc[11]

        self.assertEqual(result, t)
        
        
    def test_dataframeloader4(self):
        #Test 4: Checks whether cutting redundant data influences the data beyond removing rows
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'))
        
        t = '""Regel5'
        result = result['Header1'].loc[5]

        self.assertEqual(result, t)        


#    def test_dataframloader5(self):
#        #Test 5: Check whether the dataframe returns values rather than formulas
#        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'))        
#        result = result['Header1'].loc[2]
#        t = 18
#
#        self.assertEqual(result, t)  
        
        
    def test_dataframeloader6(self):
        #Test 6: Check whether the dataframe structure retains quoted formulas
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'))
        
        t = '=17^6'
        result = result['Header2'].loc[3]

        self.assertEqual(result, t)  


    def test_dataframeloader7(self):
        #Test 7: Checks whether slicing retains it's values
        result = self.dataframeloader.load_data(data_location = locator('dataframeloader1.xlsx'))
        
        t = [['Regel1', 'Regel2', 'Regel3'], ['"3454235', '&85674#', "'sadaf'"]]
        df = DataFrame(t, columns = ['Header1', 'Header2', 'Header3'])    
        df.index.rename('ID', inplace = True)
        
        assert_frame_equal(result.iloc[[1]], df.iloc[[1]])
        

    def test_dataframeloader8(self):
        #Test 8: Check whether the col_is_index function works properly
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'), col_is_index = 'Header1')
        result = result['Header2'].loc['Regel1']
        
        self.assertEqual(result, 'Regel2')  
        
    def test_dataframeloader9(self):
        #Test 9: Check whether changing the index changes the usage of slicing 
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'), col_is_index = 'Header1')
        Header1 = ['Regel1', '"3454235' ]        
        t = [['Regel2', 'Regel3'], ['&85674#', "'sadaf'"]]
        df = DataFrame(t, columns = ['Header2', 'Header3'], index = Header1)    
        df.index.rename('Header1', inplace = True)
            
        assert_frame_equal(result.iloc[[1]], df.iloc[[1]])          
        
    def test_dataframeloader10(self):
        #Test 10: Check whether changing the index changes the usage of slicing 
        result = self.dataframeloader.load_data(data_location = locator('datalistloader1.xlsx'), col_is_index = 'Header1')
        Header1 = ['Regel1', '"3454235' ]        
        t = [['Regel2', 'Regel3'], ['&85674#', "'sadaf'"]]
        df = DataFrame(t, columns = ['Header2', 'Header3'], index = Header1)    
        df.index.rename('Header1', inplace = True)
            
        assert_frame_equal(result.iloc[[1]], df.iloc[[1]])          
        
class DataWriterSetUpTearDown(unittest.TestCase):
    
    """
    Setup doesn't currently clear the worksheet. It should for proper usage
    """
    
#    def setUp(self):
#        wb = xl.load_workbook(locator('writetest1.xlsx'))
#        for sheetname in wb.get_sheet_names():   
#            ws = wb.get_sheet_by_name(sheetname)
#            for row in range(1, 50):
#                for col in range(1, 50):
#                    ws.cell(column = col, row = row, value=1)
#        if ws is not None:
#            wb.remove_sheet(ws)
#            
#    def tearDown(self):
#        wb = xl.load_workbook(locator('writetest1.xlsx'))
#        for sheetname in wb.get_sheet_names():   
#            ws = wb.get_sheet_by_name(sheetname)
#            for row in range(1, 50):
#                for col in range(1, 50):
#                    ws.cell(column = col+1, row = row +1, value= 0)  
                    
class DataWriterTestCase(DataWriterSetUpTearDown):
    """
    Any problems arising here might also be due to the DataListLoader. Keep that in mind when errors occur
    """
    
    """
    Might we need an option for multitab usage? I'm not sure how much we're going to use this class anyway but it's something
    to consider should we use it in the end.
    """
    datawriter = DL.DataWriter()
    dataloader = DL.DataListLoader()
    
    def test_datawriter1(self):
        #Test 1: Check whether the datawriter's basics work
        self.datawriter.append_list_to_excelfile(locator('writetest1.xlsx'), [['Regel1', 'Regel2', 'Regel3'],
                                                                             ['Regel4', 'Regel5', 'Regel6'], [None, None, None],
                                                                             ['Regel7', 'Regel8', 'Regel9']])
        result = self.dataloader.load_excel_to_lists(locator('writetest1.xlsx'))

        self.assertEqual(result[1][2], 'Regel3')
        
    def test_datawriter2(self):
        #Test 2: Check whether the datawriter writes multi-line input correctly
        self.datawriter.append_list_to_excelfile(locator('writetest1.xlsx'), [['Regel1', 'Regel2', 'Regel3'],
                                                                             ['Regel4', 'Regel5', 'Regel6']])
        result = self.dataloader.load_excel_to_lists(locator('writetest1.xlsx'))

        self.assertEqual(result[2][0], 'Regel4')    
        
    def test_datawriter3(self):
        #Test 3: Check whether the datawriter writes multi-line input correctly across the entire row
        self.datawriter.append_list_to_excelfile(locator('writetest1.xlsx'), [['Regel1', 'Regel2', 'Regel3'],
                                                                             ['Regel4', 'Regel5', 'Regel6']])
        result = self.dataloader.load_excel_to_lists(locator('writetest1.xlsx'))

        self.assertEqual(result[2][2], 'Regel6')            
        
    def test_datawriter4(self):
        #Test 4: Check whether the datawriter writes multi-line input correctly when empty rows are added
        self.datawriter.append_list_to_excelfile(locator('writetest1.xlsx'), [['Regel1', 'Regel2', 'Regel3'],
                                                                             ['Regel4', 'Regel5', 'Regel6'], [None, None, None],
                                                                             ['Regel7', 'Regel8', 'Regel9']])
        result = self.dataloader.load_excel_to_lists(locator('writetest1.xlsx'))
        
        self.assertEqual(result[4][2], 'Regel9')
        
    
    
if __name__ == '__main__':
    test_classes_to_run = [DataListLoaderTestCases, DataCutterTestCases, DataFrameTestCase, DtypeDetectorTestCase, DataWriterTestCase]

    loader = unittest.TestLoader()

    suites_list = []
    for test_class in test_classes_to_run:
        suite = loader.loadTestsFromTestCase(test_class)
        suites_list.append(suite)

    big_suite = unittest.TestSuite(suites_list)

    runner = unittest.TextTestRunner()
    results = runner.run(big_suite)






















































































































































































































































































































































































































































































































































































print ('woo')





















































































































































































































































































































































































































































































































































































































